"""Unit tests for submission verification, history sync, and the submission
service guards (docs/10). The real click boundary is covered in browser E2E."""

import json

import pytest
from openpyxl import load_workbook

from job_platform.ats.greenhouse import GreenhouseAdapter
from job_platform.browser.models import FieldType, FormField, PageSnapshot
from job_platform.candidate.loader import load_candidate_bundle
from job_platform.packages.models import PackageJobSummary
from job_platform.packages.store import PackageStore
from job_platform.preparation.service import PreparationService
from job_platform.providers.mock import MockReasoningProvider
from job_platform.shared.config import Settings
from job_platform.storage.tracker import ApplicationRecord, ApplicationTracker
from job_platform.submission.history import ApplicationHistoryService
from job_platform.submission.models import (
    AttemptStatus,
    EvidenceStrength,
    SubmissionAttempt,
    SubmissionOutcome,
)
from job_platform.submission.service import SubmissionBlockedError, SubmissionService
from job_platform.submission.verifier import (
    SubmissionVerifier,
    extract_confirmation_number,
)
from tests.unit.test_review import make_ranked

JOB = PackageJobSummary(
    company="ExampleCo", title="Backend Engineer", job_id="42",
    application_url="https://example.com/jobs/42",
)


def snap(url="https://example.com/apply", heading="", title="", text="",
         fields=0, errors=None) -> PageSnapshot:
    return PageSnapshot(
        url=url, heading=heading, title=title, text_excerpt=text,
        fields=[
            FormField(field_id=f"f{i}", selector=f"#f{i}", field_type=FieldType.TEXT)
            for i in range(fields)
        ],
        validation_errors=errors or [],
    )


BEFORE = snap(fields=6)


class TestConfirmationNumber:
    def test_extracts_labeled_number(self):
        result = extract_confirmation_number(
            "Thanks! Confirmation Number: GH-483726 has been assigned."
        )
        assert result.value == "GH-483726"
        assert result.label == "Confirmation Number"

    @pytest.mark.parametrize("label", [
        "Application Number", "Submission ID", "Reference Number",
        "Application Reference",
    ])
    def test_supported_labels(self, label):
        assert extract_confirmation_number(f"{label}: ABC-123456").value == "ABC-123456"

    def test_no_number_returns_none(self):
        assert extract_confirmation_number("Thank you for applying!") is None


class TestVerifierClassification:
    def verifier(self):
        return SubmissionVerifier()

    def test_confirmation_number_with_identity_is_conclusive_submitted(self):
        after = snap(
            url="https://example.com/confirmation",
            heading="Thank you for applying to ExampleCo",
            text="Your application has been received. Confirmation Number: AB-1",
        )
        result = self.verifier().classify("a1", BEFORE, after, JOB)
        assert result.outcome == SubmissionOutcome.SUBMITTED
        assert result.confidence >= 95
        assert result.confirmation_number.value == "AB-1"
        assert result.job_identity_verified

    def test_explicit_message_with_identity_is_strong_submitted(self):
        after = snap(
            heading="Application submitted",
            text="Backend Engineer at ExampleCo — application submitted.",
        )
        result = self.verifier().classify("a1", BEFORE, after, JOB)
        assert result.outcome == SubmissionOutcome.SUBMITTED
        assert 90 <= result.confidence < 95

    def test_generic_success_without_identity_is_unknown(self):
        after = snap(heading="Success!", text="Your application has been received.")
        result = self.verifier().classify("a1", BEFORE, after, JOB)
        assert result.outcome == SubmissionOutcome.SUBMISSION_UNKNOWN
        assert not result.job_identity_verified

    def test_weak_redirect_is_unknown(self):
        after = snap(url="https://example.com/careers", heading="Open positions")
        result = self.verifier().classify("a1", BEFORE, after, JOB)
        assert result.outcome == SubmissionOutcome.SUBMISSION_UNKNOWN
        strengths = {e.strength for e in result.evidence}
        assert strengths <= {EvidenceStrength.WEAK, EvidenceStrength.SUPPORTING}

    def test_validation_errors_mean_failed(self):
        after = snap(
            heading="Apply", fields=6, errors=["Email address is required."]
        )
        result = self.verifier().classify("a1", BEFORE, after, JOB)
        assert result.outcome == SubmissionOutcome.FAILED

    def test_already_applied_detected(self):
        after = snap(text="You have already applied to this position at ExampleCo.")
        result = self.verifier().classify("a1", BEFORE, after, JOB)
        assert result.outcome == SubmissionOutcome.ALREADY_APPLIED

    def test_closed_job_detected(self):
        after = snap(heading="This job is no longer open")
        result = self.verifier().classify("a1", BEFORE, after, JOB)
        assert result.outcome == SubmissionOutcome.APPLICATION_CLOSED

    def test_adapter_confirmation_contributes(self):
        after = snap(
            heading="Thank you for applying to ExampleCo",
            title="Job Application Confirmation",
        )
        result = SubmissionVerifier(GreenhouseAdapter()).classify("a1", BEFORE, after, JOB)
        assert result.outcome == SubmissionOutcome.SUBMITTED
        assert any(e.source == "ats_adapter" for e in result.evidence)


@pytest.fixture
def history_env(tmp_path):
    tracker = ApplicationTracker(tmp_path / "tracker.csv")
    history = ApplicationHistoryService(
        tracker, tmp_path / "history_events.jsonl", tmp_path / "tracker.xlsx"
    )
    return tracker, history, tmp_path


class TestHistoryService:
    def record(self, job_id="42", title="Backend Engineer"):
        return ApplicationRecord(
            company="ExampleCo", job_title=title, job_id=job_id,
            application_url=f"https://example.com/jobs/{job_id}", status="submitted",
        )

    def test_sync_is_idempotent(self, history_env):
        tracker, history, tmp = history_env
        assert history.sync_submission(self.record(), package_id="p1") is True
        assert history.sync_submission(self.record(), package_id="p1") is False
        assert len(tracker.records()) == 1
        types = [e.event_type for e in history.events()]
        assert "history_synced" in types
        assert "history_already_synced" in types

    def test_xlsx_rebuilt_from_csv(self, history_env):
        tracker, history, tmp = history_env
        history.sync_submission(self.record("42"))
        history.sync_submission(self.record("43", title="Platform Engineer"))
        workbook = load_workbook(tmp / "tracker.xlsx")
        sheet = workbook["Applications"]
        assert sheet.max_row == 3  # header + 2 records
        summary = workbook["Status Summary"]
        rows = {row[0].value: row[1].value for row in summary.iter_rows(min_row=2)}
        assert rows["submitted"] == 2
        assert rows["total"] == 2

    def test_corrupt_xlsx_recovered_by_rebuild(self, history_env):
        tracker, history, tmp = history_env
        history.sync_submission(self.record())
        (tmp / "tracker.xlsx").write_text("corrupt garbage")
        history.rebuild_xlsx()
        workbook = load_workbook(tmp / "tracker.xlsx")
        assert workbook["Applications"].max_row == 2

    def test_events_are_append_only_and_filterable(self, history_env):
        tracker, history, _ = history_env
        history.record_event("submitted", package_id="p1")
        history.record_event("submitted", package_id="p2")
        assert len(history.events()) == 2
        assert len(history.events(package_id="p1")) == 1


@pytest.fixture
async def submission_env(candidate_dir, tmp_path):
    bundle = load_candidate_bundle(candidate_dir)
    store = PackageStore(tmp_path / "packages")
    tracker = ApplicationTracker(tmp_path / "tracker.csv")
    history = ApplicationHistoryService(
        tracker, tmp_path / "history_events.jsonl", tmp_path / "tracker.xlsx"
    )
    prep = PreparationService(MockReasoningProvider(), store, Settings(), tracker=tracker)
    manifest = await prep.prepare(make_ranked(), bundle)
    service = SubmissionService(store, tracker, history)
    return store, tracker, history, service, manifest


class TestSubmissionGuards:
    async def test_approval_required(self, submission_env):
        *_, service, manifest = submission_env
        with pytest.raises(SubmissionBlockedError) as excinfo:
            service._guard_new_attempt(manifest, approved=False)
        assert "approval" in excinfo.value.message

    async def test_duplicate_blocks_immediately_before_submit(self, submission_env):
        store, tracker, history, service, manifest = submission_env
        tracker.add(ApplicationRecord.from_job(make_ranked().job))
        with pytest.raises(SubmissionBlockedError) as excinfo:
            service._guard_new_attempt(manifest, approved=True)
        assert "already records" in excinfo.value.message

    async def test_prior_unknown_blocks_new_attempt(self, submission_env):
        store, tracker, history, service, manifest = submission_env
        attempt = SubmissionAttempt(
            attempt_id="submission_attempt_001",
            package_id=manifest.package_id,
            status=AttemptStatus.CLICK_INITIATED,
        )
        service._persist_attempt(attempt)
        service._finalize_unknown(
            manifest, attempt, reason="browser crashed after click", missing=["page"]
        )
        with pytest.raises(SubmissionBlockedError) as excinfo:
            service._guard_new_attempt(manifest, approved=True)
        assert "Submission Unknown" in excinfo.value.message
        # Unknown state survives 'restart' (fresh service over same store)
        fresh = SubmissionService(store, tracker, history)
        assert fresh.load_unknown_outcome(manifest.package_id) is not None
        with pytest.raises(SubmissionBlockedError):
            fresh._guard_new_attempt(manifest, approved=True)

    async def test_click_initiated_without_outcome_blocks(self, submission_env):
        store, tracker, history, service, manifest = submission_env
        attempt = SubmissionAttempt(
            attempt_id="submission_attempt_001",
            package_id=manifest.package_id,
            status=AttemptStatus.CLICK_INITIATED,
        )
        service._persist_attempt(attempt)
        with pytest.raises(SubmissionBlockedError) as excinfo:
            service._guard_new_attempt(manifest, approved=True)
        assert "clicked submit" in excinfo.value.message

    async def test_submitted_package_blocks_reattempt(self, submission_env):
        store, tracker, history, service, manifest = submission_env
        attempt = SubmissionAttempt(
            attempt_id="submission_attempt_001",
            package_id=manifest.package_id,
            status=AttemptStatus.SUBMITTED,
        )
        service._persist_attempt(attempt)
        with pytest.raises(SubmissionBlockedError) as excinfo:
            service._guard_new_attempt(manifest, approved=True)
        assert "already submitted" in excinfo.value.message

    async def test_failed_before_click_allows_retry(self, submission_env):
        store, tracker, history, service, manifest = submission_env
        attempt = SubmissionAttempt(
            attempt_id="submission_attempt_001",
            package_id=manifest.package_id,
            status=AttemptStatus.FAILED_BEFORE_CLICK,
        )
        service._persist_attempt(attempt)
        service._guard_new_attempt(manifest, approved=True)  # does not raise


class TestUnknownResolution:
    async def test_resolve_to_submitted_syncs_history(self, submission_env):
        store, tracker, history, service, manifest = submission_env
        attempt = SubmissionAttempt(
            attempt_id="submission_attempt_001",
            package_id=manifest.package_id,
            status=AttemptStatus.CLICK_INITIATED,
        )
        service._persist_attempt(attempt)
        service._finalize_unknown(manifest, attempt, reason="crash", missing=["page"])

        resolution = service.resolve_unknown(
            manifest, SubmissionOutcome.SUBMITTED, "user_confirmation",
            notes="found it in my email",
        )
        assert resolution.resolved_status == SubmissionOutcome.SUBMITTED
        assert service.load_unknown_outcome(manifest.package_id) is None
        assert len(tracker.records()) == 1
        # Attempt updated, lock released, resolution recorded
        attempts = service.load_attempts(manifest.package_id)
        assert attempts[0].status == AttemptStatus.SUBMITTED
        assert not (store.package_dir(manifest.package_id) / "submission/.submission.lock").exists()
        resolution_file = json.loads(
            (store.package_dir(manifest.package_id) / "submission/unknown_resolution.json").read_text()
        )
        assert resolution_file["resolution_source"] == "user_confirmation"

    async def test_resolve_to_failed_allows_future_attempt(self, submission_env):
        store, tracker, history, service, manifest = submission_env
        attempt = SubmissionAttempt(
            attempt_id="submission_attempt_001",
            package_id=manifest.package_id,
            status=AttemptStatus.CLICK_INITIATED,
        )
        service._persist_attempt(attempt)
        service._finalize_unknown(manifest, attempt, reason="crash", missing=["page"])
        service.resolve_unknown(manifest, SubmissionOutcome.FAILED, "ats_dashboard")
        service._guard_new_attempt(manifest, approved=True)  # retry now allowed

    async def test_resolution_without_unknown_raises(self, submission_env):
        *_, service, manifest = submission_env
        from job_platform.shared.errors import StorageError

        with pytest.raises(StorageError):
            service.resolve_unknown(manifest, SubmissionOutcome.FAILED, "user")
